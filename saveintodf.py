import os
import openpyxl
import datetime
import pandas as pd
import win32com.client as win32
import warnings
import py7zr

warnings.filterwarnings("ignore")

def extract_formulas(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
    formulas_with_details = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and str(cell.value).startswith('='):
                    formula = str(cell.value)                    
                    coordinates = cell.coordinate
                    formulas_with_details.append((file_path, os.path.basename(file_path), sheet, coordinates, formula.replace('=','')))
    df = pd.DataFrame(formulas_with_details, columns=['Path', 'Excel Name', 'Sheet Name', 'Coordinates', 'Formulas'])
    return df

def extract_format(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
    formats_with_details = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if not ws.sheet_state == 'hidden':
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        font = cell.font
                        fill = cell.fill
                        border = cell.border
                        alignment = cell.alignment
                        coordinates = cell.coordinate
                        font_details = f"Font Name: {font.name}, Size: {font.sz}, Color: {font.color}, Bold: {font.b}, Italic: {font.i}"
                        format_details = f"{font_details}, Fill: {fill}, Border: {border}, Alignment: {alignment}"
                        formats_with_details.append((file_path, os.path.basename(file_path), sheet, coordinates, format_details))  # Append the necessary details
    df = pd.DataFrame(formats_with_details, columns=['Path', 'Excel Name', 'Sheet Name', 'Coordinates', 'Format Details'])  # Update the column names
    return df

countofformats = 0
def compare_formats(df1,df2):
    global countofformats
    merged_df = pd.merge(df1, df2, on=['Sheet Name', 'Coordinates'], how='outer', suffixes=('_1', '_2'))
    merged_df['Status'] = merged_df['Format Details_1'] == merged_df['Format Details_2']
    countofformats = len(merged_df[merged_df['Status'] == False])
    return merged_df

countofflase = 0

def compare_formulas(df1, df2):
    global countofflase
    merged_df = pd.merge(df1, df2, on=['Sheet Name', 'Coordinates'], how='outer', suffixes=('_1', '_2'))
    merged_df['Status'] = merged_df['Formulas_1'] == merged_df['Formulas_2']
    countofflase = len(merged_df[merged_df['Status'] == False])
    return merged_df

def sendmail(recipients, file_path3, file_path4, reportname=None, Location1=None, Location2=None, cc_recipients=None):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipients
    # mail.CC = cc_recipients

    files_to_compress = ["formats_comparison.xlsx", "formulas_comparison.xlsx"]

    with py7zr.SevenZipFile("Comparission_files.7z", "w") as archive:  # Faster, less compression
        for filename in files_to_compress:
            archive.write(filename)
    for acc in outlook.Session.Accounts:
        if str(acc) == 'saimanojb@maqsoftware.com':
            reqmail = acc
            break
    mail._oleobj_.Invoke(*(64209, 0, 8, 0, reqmail))  # Set the sender account
    status = 'Success' if countofflase == 0 else 'Failure'
    mail.Subject = status + ' | ' + 'Excel Formula Comparison' + ' | ' + reportname
    attachment = os.getcwd() + '\\Comparission_files.7z'
    mail.Attachments.Add(attachment)
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    html_body = f'''
    <html>
    <body>
        <h2>Excel Formula Comparison Report</h2>
        <p><strong>Report Name:</strong> {reportname}</p>
        <p><strong>Location 1:</strong> {Location1}</p>
        <p><strong>Location 2:</strong> {Location2}</p>
        <p><strong>Status:</strong> <span style="color: {'green' if status == 'Success' else 'red'}"><strong>{status}</strong></span></p>
        <p><strong>No of not matching formulas:</strong> {countofflase}</p>
        <p><strong>No of not matching formats:</strong> {countofformats}</p>
        <p><strong>Executed time:</strong> {current_time}</p>
        <br>
        <p>Thanks and Regards,</p>
        <p><i>Reporting Dev Team</i></p>
    </body>
    </html>
    '''
    mail.HTMLBody = html_body
    mail.Send()
    print("Mail sent successfully")

def conditional_formatting(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=True)
    conditional_formatting_rules = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if not ws.sheet_state == 'hidden':
            for range_ in ws.conditional_formatting:
                for rule in ws.conditional_formatting[range_]:
                    # Extract range, type, formula, and detailed format
                    rule_range = range_.sqref
                    rule_type = rule.type
                    rule_formula = rule.formula
                    # Extract detailed formatting information
                    bg_color = (
                        rule.dxf.fill.bgColor.rgb
                        if rule.dxf.fill and rule.dxf.fill.bgColor
                        else None
                    )
                    font_style = {
                        'bold': rule.dxf.font.b,
                        'italic': rule.dxf.font.i,
                        # Add more font attributes as needed
                    }

                    border_style = {
                        # Extract border attributes as needed
                    }

                    # Convert RGB object to a string representation
                    bg_color_str = str(bg_color) if bg_color else None

                    # Create a dictionary for each rule
                    rule_data = {
                        'RANGE': rule_range,
                        # 'type': rule_type,
                        'formula': rule_formula,
                        'format': {
                            'bg_color': bg_color_str,
                            'font_style': font_style,
                            'border_style': border_style,
                            # Add more formatting categories as needed
                        },
                    }
                    conditional_formatting_rules.append(rule_data)
    df = pd.DataFrame(conditional_formatting_rules, columns=['RANGE', 'Formula', 'Format'])
    return df

def compare_cformats(df1, df2):
    merged_df = pd.merge(df1, df2, on=['RANGE'], how='outer', suffixes=('_1', '_2'))
    merged_df['Status'] = merged_df['Format_1'] == merged_df['Format_2']
    return merged_df
                
if __name__ == '__main__':
    file_path1 = 'C:\\Users\\v-mbnvsai\\OneDrive - Microsoft\\Desktop\\Dev\\American Triangulation - January Forecast Dev.xlsm'
    file_path2 = 'C:\\Users\\v-mbnvsai\\OneDrive - Microsoft\\Desktop\\Prod\\American Triangulation - January Forecast Prod.xlsm'
    reportname = os.path.basename(file_path1)

    Location1 = os.path.basename(os.path.dirname(file_path1))
    Location2 = os.path.basename(os.path.dirname(file_path2))
    df1 = extract_formulas(file_path1)
    df2 = extract_formulas(file_path2)
    merged_df = compare_formulas(df1, df2)
    merged_df.to_excel('formulas_comparison.xlsx', index=False)
    file_path3 = 'C:\\Users\\v-mbnvsai\\OneDrive - Microsoft\\Documents\\Learning\\Py\\Formulas\\formulas_comparison.xlsx'
    # recipients = 'seshuc@maqsoftware.com'
    # #cc_recipients = 'saipavanm@maqsoftware.com'
    recipients = input("Enter recipients: ")
    cc_recipients = input("Enter cc recipients: ")

    df3 = extract_format(file_path1)
    df4 = extract_format(file_path2)
    merged_df1 = compare_formats(df3, df4)
    merged_df1.to_excel('formats_comparison.xlsx', index=False)
    file_path4 = 'C:\\Users\\v-mbnvsai\\OneDrive - Microsoft\\Documents\\Learning\\Py\\Formulas\\formats_comparison.xlsx'
    # df5 = conditional_formatting(file_path1)
    # df6 = conditional_formatting(file_path2)
    # merged_df2 = compare_cformats(df5, df6)
    # merged_df2.to_excel('cformats_comparison.xlsx', index=False)
    sendmail(recipients, file_path3, file_path4, reportname, Location1, Location2, cc_recipients=None)